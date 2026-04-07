from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from functools import lru_cache
from io import BytesIO
from pathlib import Path
import threading
from typing import Literal
from uuid import uuid4

from fastapi import FastAPI, File, Form, HTTPException, Query, UploadFile
from fastapi.responses import FileResponse, StreamingResponse
from openpyxl import load_workbook
from pydantic import BaseModel, Field

from app.db import get_conn

app = FastAPI(title="MAKS Reverse Geocoding", version="0.5.0")
UI_PATH = Path(__file__).resolve().parent / "ui" / "index.html"
MAX_BATCH_POINTS = 2000
MAX_PARALLEL_WORKERS = 8
MAX_COORD_TEXT_LEN = 64
EXCEL_JOBS: dict[str, dict] = {}
EXCEL_JOBS_LOCK = threading.Lock()
EXCEL_HEADERS = [
    "IL",
    "ILCE",
    "KOY",
    "KOY_BULMA_YONTEMI",
    "MAHALLE",
    "MAHALLE_BULMA_YONTEMI",
    "EN_YAKIN_CADDE_SOKAK",
    "BINADAN_GELEN_CADDE_SOKAK",
    "BINA_NO",
    "KAPI_NO",
]


class BatchPoint(BaseModel):
    id: str | None = None
    lat: float = Field(..., ge=-90, le=90)
    lon: float = Field(..., ge=-180, le=180)


class BatchReverseRequest(BaseModel):
    points: list[BatchPoint] = Field(..., min_length=1, max_length=MAX_BATCH_POINTS)
    door_radius_m: float = Field(20.0, gt=0, le=500)
    building_radius_m: float = Field(60.0, gt=0, le=1000)
    road_radius_m: float = Field(120.0, gt=0, le=2000)
    metric: Literal["geodesic", "planar"] = "geodesic"
    parallel_workers: int = Field(1, ge=1, le=MAX_PARALLEL_WORKERS)


@app.get("/health")
def health() -> dict:
    return {"status": "ok"}


@app.get("/")
def ui_root() -> FileResponse:
    return FileResponse(UI_PATH)


@app.get("/ui")
def ui_page() -> FileResponse:
    return FileResponse(UI_PATH)


def _build_query(metric: str) -> tuple[str, str]:
    pt_expr = "ST_SetSRID(ST_MakePoint(%s, %s), 4326)"
    norm_tpl = "UPPER(REGEXP_REPLACE(BTRIM(CAST({expr} AS text)), '\\.0+$', ''))"

    if metric == "geodesic":
        dist = "ST_Distance(p.geom::geography, {geom}::geography)"
        within = "ST_DWithin(p.geom::geography, {geom}::geography, %s)"
    else:
        dist = "ST_Distance(ST_Transform(p.geom, 3857), ST_Transform({geom}, 3857))"
        within = "ST_DWithin(ST_Transform(p.geom, 3857), ST_Transform({geom}, 3857), %s)"

    query = """
        WITH p AS (
          SELECT {pt_expr} AS geom
        ),
        admin AS (
          SELECT
            il_hit.ad AS il,
            ilce_hit.ad AS ilce,
            koy_hit.ad AS koy,
            CASE
              WHEN koy_hit.is_inside THEN 'poligon_icinde'
              ELSE 'en_yakin'
            END AS koy_bulma_yontemi,
            mahalle_hit.ad AS mahalle,
            CASE
              WHEN mahalle_hit.is_inside THEN 'poligon_icinde'
              ELSE 'en_yakin'
            END AS mahalle_bulma_yontemi
          FROM p
          LEFT JOIN LATERAL (
            SELECT i.ad
            FROM raw_maks.il i
            ORDER BY ST_Covers(i.geom, p.geom) DESC, i.geom <-> p.geom
            LIMIT 1
          ) il_hit ON TRUE
          LEFT JOIN LATERAL (
            SELECT d.ad
            FROM raw_maks.ilce d
            ORDER BY ST_Covers(d.geom, p.geom) DESC, d.geom <-> p.geom
            LIMIT 1
          ) ilce_hit ON TRUE
          LEFT JOIN LATERAL (
            SELECT k.ad
                 , ST_Covers(k.geom, p.geom) AS is_inside
            FROM raw_maks.koy k
            ORDER BY ST_Covers(k.geom, p.geom) DESC, k.geom <-> p.geom
            LIMIT 1
          ) koy_hit ON TRUE
          LEFT JOIN LATERAL (
            SELECT m.ad
                 , ST_Covers(m.geom, p.geom) AS is_inside
            FROM raw_maks.mahalle m
            ORDER BY ST_Covers(m.geom, p.geom) DESC, m.geom <-> p.geom
            LIMIT 1
          ) mahalle_hit ON TRUE
          LIMIT 1
        ),
        nearest_door AS (
          SELECT
            nm.id AS door_id,
            nm.kapino AS kapi_no,
            nm.yapiid AS yapi_id,
            {door_dist} AS dist_m
          FROM raw_maks.numarataj nm, p
          WHERE {door_within}
          ORDER BY nm.geom <-> p.geom
          LIMIT 1
        ),
        nearest_yapi AS (
          SELECT
            y.id AS yapi_id,
            y.ad AS bina_no,
            {building_dist} AS dist_m
          FROM raw_maks.yapi y, p
          WHERE {building_within}
          ORDER BY y.geom <-> p.geom
          LIMIT 1
        ),
        selected_yapi AS (
          SELECT
            COALESCE(nd.yapi_id, ny.yapi_id) AS yapi_id,
            ny.bina_no,
            ny.dist_m AS yapi_dist_m
          FROM nearest_door nd
          FULL OUTER JOIN nearest_yapi ny ON TRUE
          LIMIT 1
        ),
        yapi_road AS (
          SELECT
            yol.ad AS road_name,
            COUNT(*) AS cnt
          FROM selected_yapi sy
          JOIN raw_maks.numarataj nm
            ON {norm_nm_yapi} = {norm_sy_yapi}
          JOIN raw_maks.yolortahatyon yhy
            ON {norm_yhy_id} = {norm_nm_yhy}
          JOIN raw_maks.yolortahat yoh
            ON {norm_yoh_id} = {norm_yhy_yoh}
          JOIN raw_maks.yol yol
            ON {norm_yol_id} = {norm_yoh_yol}
          GROUP BY yol.ad
          ORDER BY COUNT(*) DESC, yol.ad
          LIMIT 1
        ),
        nearest_road AS (
          SELECT
            yol.ad AS road_name,
            {road_dist} AS dist_m
          FROM raw_maks.yolortahat yoh
          CROSS JOIN p
          LEFT JOIN raw_maks.yol yol
            ON {norm_yol_id} = {norm_yoh_yol}
          WHERE {road_within}
          ORDER BY yoh.geom <-> p.geom
          LIMIT 1
        )
        SELECT
          a.il,
          a.ilce,
          a.koy,
          a.koy_bulma_yontemi,
          a.mahalle,
          a.mahalle_bulma_yontemi,
          nd.kapi_no,
          sy.bina_no,
          nr.road_name AS en_yakin_cadde_sokak,
          yr.road_name AS yapinin_bagli_oldugu_cadde_sokak,
          nd.dist_m AS door_dist_m,
          sy.yapi_dist_m AS building_dist_m,
          nr.dist_m AS road_dist_m,
          CASE
            WHEN nd.door_id IS NOT NULL THEN 'door'
            WHEN sy.yapi_id IS NOT NULL THEN 'building'
            WHEN nr.road_name IS NOT NULL THEN 'road'
            ELSE 'none'
          END AS source_level
        FROM admin a
        LEFT JOIN nearest_door nd ON TRUE
        LEFT JOIN selected_yapi sy ON TRUE
        LEFT JOIN yapi_road yr ON TRUE
        LEFT JOIN nearest_road nr ON TRUE
        LIMIT 1
        """.format(
        pt_expr=pt_expr,
        norm_nm_yapi=norm_tpl.format(expr="nm.yapiid"),
        norm_sy_yapi=norm_tpl.format(expr="sy.yapi_id"),
        norm_nm_yhy=norm_tpl.format(expr="nm.yolortahatyonid"),
        norm_yhy_id=norm_tpl.format(expr="yhy.id"),
        norm_yhy_yoh=norm_tpl.format(expr="yhy.yolortahatid"),
        norm_yoh_id=norm_tpl.format(expr="yoh.id"),
        norm_yoh_yol=norm_tpl.format(expr="yoh.yolid"),
        norm_yol_id=norm_tpl.format(expr="yol.id"),
        door_dist=dist.format(geom="nm.geom"),
        door_within=within.format(geom="nm.geom"),
        building_dist=dist.format(geom="y.geom"),
        building_within=within.format(geom="y.geom"),
        road_dist=dist.format(geom="yoh.geom"),
        road_within=within.format(geom="yoh.geom"),
    )
    return query, pt_expr


@lru_cache(maxsize=2)
def _get_cached_query(metric: str) -> str:
    query, _ = _build_query(metric)
    return query


def _row_to_payload(row: tuple, door_radius_m: float, building_radius_m: float, road_radius_m: float, metric: str) -> dict:
    (
        il,
        ilce,
        koy,
        koy_bulma_yontemi,
        mahalle,
        mahalle_bulma_yontemi,
        kapi_no,
        bina_no,
        en_yakin_cadde_sokak,
        yapinin_bagli_oldugu_cadde_sokak,
        d_d,
        b_d,
        r_d,
        source_level,
    ) = row

    yerlesim = mahalle or koy
    cadde_for_adres = yapinin_bagli_oldugu_cadde_sokak or en_yakin_cadde_sokak
    parts = [p for p in [yerlesim, cadde_for_adres, bina_no, ilce, il] if p]
    adres = ", ".join(parts) if parts else None

    confidence = 0.2
    if source_level == "door":
        confidence = 0.95
    elif source_level == "building":
        confidence = 0.8
    elif source_level == "road":
        confidence = 0.6

    return {
        "il": il,
        "ilce": ilce,
        "koy": koy,
        "koy_bulma_yontemi": koy_bulma_yontemi,
        "mahalle": mahalle,
        "mahalle_bulma_yontemi": mahalle_bulma_yontemi,
        "En yakin Cadde/Sokak": en_yakin_cadde_sokak,
        "Yapinin bagli oldugu Cadde/Sokak": yapinin_bagli_oldugu_cadde_sokak,
        "bina_no": bina_no,
        "kapi_no": kapi_no,
        "adres": adres,
        "source_level": source_level,
        "confidence": confidence,
        "distance_m": {
            "door": d_d,
            "building": b_d,
            "road": r_d,
        },
        "query_params": {
            "door_radius_m": door_radius_m,
            "building_radius_m": building_radius_m,
            "road_radius_m": road_radius_m,
            "metric": metric,
        },
    }


def _resolve_one(lat: float, lon: float, door_radius_m: float, building_radius_m: float, road_radius_m: float, metric: str) -> dict:
    query = _get_cached_query(metric)
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                query,
                (
                    lon,
                    lat,
                    door_radius_m,
                    building_radius_m,
                    road_radius_m,
                ),
            )
            row = cur.fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="Adres bulunamadi")
            return _row_to_payload(row, door_radius_m, building_radius_m, road_radius_m, metric)


def _parse_coord(value: object) -> float | None:
    if value is None:
        return None

    text = str(value).strip()
    if not text:
        return None

    if len(text) > MAX_COORD_TEXT_LEN:
        text = text[:MAX_COORD_TEXT_LEN]

    text = text.replace(" ", "")
    if "," in text and "." not in text:
        text = text.replace(",", ".")
    elif "," in text and "." in text:
        text = text.replace(",", "")

    try:
        return float(text)
    except (TypeError, ValueError):
        return None


def _try_resolve_pair(
    lat: float,
    lon: float,
    door_radius_m: float,
    building_radius_m: float,
    road_radius_m: float,
    metric: str,
) -> dict | None:
    try:
        return _resolve_one(lat, lon, door_radius_m, building_radius_m, road_radius_m, metric)
    except Exception:
        return None


def _resolve_with_coord_order(
    x: float,
    y: float,
    coord_order: str,
    door_radius_m: float,
    building_radius_m: float,
    road_radius_m: float,
    metric: str,
) -> tuple[dict | None, str]:
    mode = coord_order.lower()

    if mode == "xy":
        if not (-180 <= x <= 180 and -90 <= y <= 90):
            return None, "xy"
        return _try_resolve_pair(y, x, door_radius_m, building_radius_m, road_radius_m, metric), "xy"

    if mode == "yx":
        if not (-180 <= y <= 180 and -90 <= x <= 90):
            return None, "yx"
        return _try_resolve_pair(x, y, door_radius_m, building_radius_m, road_radius_m, metric), "yx"

    # auto: test both valid interpretations and select better confidence.
    candidates: list[tuple[str, dict]] = []

    if -180 <= x <= 180 and -90 <= y <= 90:
        res_xy = _try_resolve_pair(y, x, door_radius_m, building_radius_m, road_radius_m, metric)
        if res_xy:
            candidates.append(("xy", res_xy))

    if -180 <= y <= 180 and -90 <= x <= 90:
        res_yx = _try_resolve_pair(x, y, door_radius_m, building_radius_m, road_radius_m, metric)
        if res_yx:
            candidates.append(("yx", res_yx))

    if not candidates:
        return None, "auto"

    order, best = max(candidates, key=lambda item: float(item[1].get("confidence") or 0.0))
    return best, order


def _resolve_excel_row(
    row_idx: int,
    x: float | None,
    y: float | None,
    coord_order: str,
    door_radius_m: float,
    building_radius_m: float,
    road_radius_m: float,
    metric: str,
) -> tuple[int, dict | None, str]:
    if x is None or y is None:
        return row_idx, None, "HATA: CBS_X/CBS_Y gecersiz"

    result, used_order = _resolve_with_coord_order(
        x,
        y,
        coord_order,
        door_radius_m,
        building_radius_m,
        road_radius_m,
        metric,
    )

    if not result:
        return row_idx, None, "HATA: Koordinat cozumlenemedi (xy/yx aralik kontrolu veya sorgu sonucu)"

    result["koord_duzeni"] = used_order
    return row_idx, result, ""


def _set_job(job_id: str, **kwargs) -> None:
    with EXCEL_JOBS_LOCK:
        if job_id in EXCEL_JOBS:
            EXCEL_JOBS[job_id].update(kwargs)


def _write_excel_output(ws, row_idx: int, result: dict | None, error_text: str) -> None:
    if error_text:
        ws.cell(row=row_idx, column=4, value=error_text)
        for c in range(5, 4 + len(EXCEL_HEADERS)):
            ws.cell(row=row_idx, column=c, value="")
        return

    row_values = [
        result.get("il"),
        result.get("ilce"),
        result.get("koy"),
        result.get("koy_bulma_yontemi"),
        result.get("mahalle"),
        result.get("mahalle_bulma_yontemi"),
        result.get("En yakin Cadde/Sokak"),
        result.get("Yapinin bagli oldugu Cadde/Sokak"),
        result.get("bina_no"),
        result.get("kapi_no"),
    ]

    for i, value in enumerate(row_values, start=4):
        ws.cell(row=row_idx, column=i, value=value)


def _process_excel_job(
    job_id: str,
    file_bytes: bytes,
    door_radius_m: float,
    building_radius_m: float,
    road_radius_m: float,
    metric: str,
    parallel_workers: int,
    coord_order: str,
) -> None:
    try:
        wb = load_workbook(BytesIO(file_bytes))
        ws = wb.active

        for i, h in enumerate(EXCEL_HEADERS, start=4):
            ws.cell(row=1, column=i, value=h)

        jobs: list[tuple[int, float | None, float | None]] = []
        for r in range(2, ws.max_row + 1):
            x = _parse_coord(ws.cell(row=r, column=2).value)
            y = _parse_coord(ws.cell(row=r, column=3).value)
            jobs.append((r, x, y))

        total = len(jobs)
        _set_job(job_id, total=total, processed=0, status="running")

        worker_count = max(1, min(MAX_PARALLEL_WORKERS, int(parallel_workers)))
        row_results: list[tuple[int, dict | None, str]] = []

        if worker_count > 1:
            with ThreadPoolExecutor(max_workers=worker_count) as pool:
                fut_map = {
                    pool.submit(
                        _resolve_excel_row,
                        r,
                        x,
                        y,
                        coord_order,
                        door_radius_m,
                        building_radius_m,
                        road_radius_m,
                        metric,
                    ): r
                    for r, x, y in jobs
                }
                processed = 0
                for fut in as_completed(fut_map):
                    row_results.append(fut.result())
                    processed += 1
                    _set_job(job_id, processed=processed)
        else:
            processed = 0
            for r, x, y in jobs:
                row_results.append(
                    _resolve_excel_row(
                        r,
                        x,
                        y,
                        coord_order,
                        door_radius_m,
                        building_radius_m,
                        road_radius_m,
                        metric,
                    )
                )
                processed += 1
                _set_job(job_id, processed=processed)

        for row_idx, res, err in row_results:
            _write_excel_output(ws, row_idx, res, err)

        out = BytesIO()
        wb.save(out)
        out.seek(0)

        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_name = f"reverse_geocode_result_{stamp}.xlsx"

        _set_job(
            job_id,
            status="completed",
            output_name=output_name,
            output_bytes=out.getvalue(),
            finished_at=datetime.now().isoformat(),
        )
    except Exception as ex:
        _set_job(job_id, status="failed", error=str(ex), finished_at=datetime.now().isoformat())


@app.get("/reverse-geocode")
def reverse_geocode(
    lat: float = Query(..., ge=-90, le=90),
    lon: float = Query(..., ge=-180, le=180),
    door_radius_m: float = Query(20.0, gt=0, le=500),
    building_radius_m: float = Query(60.0, gt=0, le=1000),
    road_radius_m: float = Query(120.0, gt=0, le=2000),
    metric: Literal["geodesic", "planar"] = Query("geodesic"),
) -> dict:
    return _resolve_one(lat, lon, door_radius_m, building_radius_m, road_radius_m, metric)


@app.post("/reverse-geocode/batch")
def reverse_geocode_batch(req: BatchReverseRequest) -> dict:
    if len(req.points) > MAX_BATCH_POINTS:
        raise HTTPException(status_code=400, detail=f"Maksimum nokta sayisi {MAX_BATCH_POINTS}")

    def worker(point: BatchPoint) -> dict:
        try:
            result = _resolve_one(
                point.lat,
                point.lon,
                req.door_radius_m,
                req.building_radius_m,
                req.road_radius_m,
                req.metric,
            )
            return {
                "id": point.id,
                "lat": point.lat,
                "lon": point.lon,
                "ok": True,
                "result": result,
            }
        except Exception as ex:
            return {
                "id": point.id,
                "lat": point.lat,
                "lon": point.lon,
                "ok": False,
                "error": str(ex),
            }

    if req.parallel_workers > 1:
        with ThreadPoolExecutor(max_workers=req.parallel_workers) as pool:
            items = list(pool.map(worker, req.points))
    else:
        items = [worker(p) for p in req.points]

    success = sum(1 for i in items if i["ok"])
    failed = len(items) - success
    return {
        "count": len(items),
        "success": success,
        "failed": failed,
        "parallel_workers": req.parallel_workers,
        "query_params": {
            "door_radius_m": req.door_radius_m,
            "building_radius_m": req.building_radius_m,
            "road_radius_m": req.road_radius_m,
            "metric": req.metric,
        },
        "items": items,
    }


@app.post("/reverse-geocode/excel")
def reverse_geocode_excel_start(
    file: UploadFile = File(...),
    door_radius_m: float = Form(20.0),
    building_radius_m: float = Form(60.0),
    road_radius_m: float = Form(120.0),
    metric: Literal["geodesic", "planar"] = Form("geodesic"),
    parallel_workers: int = Form(4),
    coord_order: Literal["auto", "xy", "yx"] = Form("auto"),
) -> dict:
    filename = file.filename or "input.xlsx"
    if not filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Sadece .xlsx dosyasi destekleniyor")

    file_bytes = file.file.read()
    if not file_bytes:
        raise HTTPException(status_code=400, detail="Excel dosyasi bos")

    job_id = uuid4().hex
    with EXCEL_JOBS_LOCK:
        EXCEL_JOBS[job_id] = {
            "job_id": job_id,
            "status": "queued",
            "processed": 0,
            "total": 0,
            "created_at": datetime.now().isoformat(),
            "output_name": None,
            "output_bytes": None,
            "error": None,
            "finished_at": None,
        }

    thread = threading.Thread(
        target=_process_excel_job,
        args=(
            job_id,
            file_bytes,
            door_radius_m,
            building_radius_m,
            road_radius_m,
            metric,
            parallel_workers,
            coord_order,
        ),
        daemon=True,
    )
    thread.start()

    return {"job_id": job_id, "status": "queued"}


@app.get("/reverse-geocode/excel/status/{job_id}")
def reverse_geocode_excel_status(job_id: str) -> dict:
    with EXCEL_JOBS_LOCK:
        job = EXCEL_JOBS.get(job_id)
        if not job:
            raise HTTPException(status_code=404, detail="Job bulunamadi")

        total = int(job.get("total") or 0)
        processed = int(job.get("processed") or 0)
        percent = round((processed / total) * 100, 2) if total > 0 else 0.0
        return {
            "job_id": job_id,
            "status": job.get("status"),
            "processed": processed,
            "total": total,
            "percent": percent,
            "download_ready": job.get("status") == "completed",
            "error": job.get("error"),
        }


@app.get("/reverse-geocode/excel/download/{job_id}")
def reverse_geocode_excel_download(job_id: str) -> StreamingResponse:
    with EXCEL_JOBS_LOCK:
        job = EXCEL_JOBS.get(job_id)
        if not job:
            raise HTTPException(status_code=404, detail="Job bulunamadi")
        if job.get("status") != "completed":
            raise HTTPException(status_code=409, detail="Dosya henuz hazir degil")

        output_bytes = job.get("output_bytes")
        output_name = job.get("output_name") or "reverse_geocode_result.xlsx"

    headers = {"Content-Disposition": f'attachment; filename="{output_name}"'}
    return StreamingResponse(
        BytesIO(output_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )
